''' Open .XLSX file saved by Crystal Reports
containing order, delivery, quantity and
balance data. Search and record part numbers
showing negative balance.
Release Date: 11/11/2018. Programmer: Dave Jennings
Version 2.0 11/16/2018 changed input from .csv to
.xlsx'''

import tkinter as tk
from tkinter import filedialog  # noqa: F401
import re
import os
import xlrd
from openpyxl import Workbook
# from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle


class Purchasing_Exceptions(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.title('Purchasing Exceptions')
        self.config(bg='light blue')

        # --------------------Top Frame

        title_frame = tk.Frame(self, bd=10, bg='light blue')
        title_frame.pack(side=tk.TOP)
        title = tk.Label(title_frame,
                         text='Purchasing Exception Report Generator',
                         bg='light blue', font=('arial', 12, 'underline'))
        title.grid(column=0, row=0, sticky=tk.E+tk.W)

        # --------------------Bottom Frame

        bottom_frame = tk.Frame(self, bd=10, bg='light blue')
        bottom_frame.pack(side=tk.TOP, fill=tk.X)

        # --------------------Listbox

        self.file_listbox = tk.Listbox(bottom_frame, bg='snow', height=7,
                                       width=65, bd=4)
        self.file_listbox.grid(column=0, row=0, columnspan=4,
                               rowspan=3, sticky=tk.E+tk.W)

        text_scroll = tk.Scrollbar(bottom_frame, orient=tk.VERTICAL,
                                   command=self.file_listbox.yview)
        self.file_listbox['yscroll'] = text_scroll.set
        text_scroll.place(in_=self.file_listbox, relx=.96, relheight=1.0,
                          bordermode='inside')

        # --------------------Buttons

        self.choose_file_button = tk.Button(bottom_frame, text="Select File",
                                            relief=tk.RAISED, width=16, bd=4,
                                            padx=10, pady=6, bg='snow3')
        self.choose_file_button.bind('<ButtonRelease-1>', self.choose_files)
        self.choose_file_button.grid(column=5, row=0)

        self.process_file_button = tk.Button(bottom_frame, text="Process",
                                             relief=tk.RAISED, width=16, bd=4,
                                             padx=10, pady=6, bg='snow3')
        self.process_file_button.grid(column=5, row=1)

        self.open_results_button = tk.Button(bottom_frame,
                                             text="Open Results File",
                                             relief=tk.RAISED, width=16, bd=4,
                                             padx=10, pady=6, bg='snow3')
        self.open_results_button.grid(column=5, row=2)

        # --------------------Buttons

        menubar = tk.Menu(self)
        self.config(menu=menubar)
        sub_menu = tk.Menu(menubar, tearoff=False)
        menubar.add_cascade(label='File', menu=sub_menu)
        sub_menu.add_command(label='Open', command=self.open_file)

    def choose_files(self, event):
        ''' Choose input file'''
        self.filename = (tk.filedialog.askopenfilename(title='Add File',
                         filetypes=(('Excel Files', '*.xlsx'),
                                    ('All Files', '*.*'))))
        self.idir = os.path.split(self.filename)[0]
        self.file_listbox.insert(tk.END, self.filename)
        # self.files.append(self.filename)
        self.process_file_button.bind('<ButtonRelease-1>', self.run_analysis)

    def run_analysis(self, event):
        ''' Search input file for part # and negative balance'''
        ds = set()
        wb = xlrd.open_workbook(self.filename)
        sht = wb.sheet_by_index(0)
        ds = set()
        for row in range(0, sht.nrows):
            x = sht.row_values(row)
            pnum_match = re.search(r'Part #:', x[0])
            if pnum_match:
                pnum = x[1]
            elif str(x[6]).startswith('-'):
                ds.add(pnum)
                ds_list = list(ds)
                fin = sorted(ds_list)
        self.part_list = []
        for item in fin:
            self.part_list.append(item)
            self.file_listbox.insert(tk.END, item)
        self.count_string = (('Processing complete. {} files found.')
                             .format(len(fin)))
        self.file_listbox.see(tk.END)
        self.open_results_button.bind('<ButtonRelease-1>',
                                      self.open_results_file)
        self.choose_file_button.unbind('<ButtonRelease-1>')
        self.process_file_button.unbind('<ButtonRelease-1>')
        self.write_to_spreadsheet()

    def write_to_spreadsheet(self):
        '''Formats and writes data to spreadsheet.'''
        wb = Workbook()
        highlight = NamedStyle(name="highlight")
        highlight.font = Font(bold=False, size=11)
        bd = Side(style='thin', color="000000")
        highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        wb.add_named_style(highlight)  # Register named style
        sh1 = wb.active
        sh1.title = 'Purchasing Exceptions'
        sh1.append(['Part Number'])
        sh1['A1'].font = Font(bold=True, size=11)
        sh1['A1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        sh1['A1'].alignment = Alignment(horizontal='center')
        rnum = 2
        sh1.column_dimensions['A'].width = (30)
        # col_width2 = 0
        for item in self.part_list:
            # sh1_tool_list_data = item
            sh1.cell(row=rnum, column=1).value = item
            sh1.cell(row=rnum, column=1).style = 'highlight'
            # if len(str(sh1_description)) > col_width2:
            #     col_width2 = len(str(sh1_description))
            rnum += 1
        # sh1.column_dimensions['A'].width = (col_width2 * 1.125)
        save_name = (('{}/Purchasing Exception Report.xlsx')
                     .format(self.idir))

        self.saved_as_string = ('Results file: {}').format(save_name)
        self.file_listbox.insert(tk.END, self.count_string)
        self.file_listbox.insert(tk.END, self.saved_as_string)
        self.file_listbox.see(tk.END)
        wb.save(save_name)
        os.startfile(save_name)

    def open_results_file(self, event):
        """ Open results file in Excel"""
        os.startfile('Purchasing Exception Report.xlsx')

    def open_file(self):
        ''' Open any file'''
        open_filename = (tk.filedialog.askopenfilename(title='Add File',
                         filetypes=(('Excel Files', '*.xlsx'),
                                    ('All Files', '*.*'))))
        os.startfile(open_filename)


root = Purchasing_Exceptions()
root.mainloop()
